from __future__ import annotations

import csv
import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from pypdf import PdfReader

try:
    from openpyxl import load_workbook
except Exception:  # pragma: no cover
    load_workbook = None


SOURCE_ROOT = Path("/Users/tomasz/Documents/Content Elevated/Content Elevated Products")
WORKSPACE_ROOT = Path("/Users/tomasz/Documents/Codex/2026-05-15/can-you-build-a-website-for")
OUT_DIR = WORKSPACE_ROOT / "workspace-assets" / "product-audit"


CUSTOMER_SKIP = {"_SELLER MARKETING & INTERNAL DOCS"}
SPREADSHEET_EXTENSIONS = {".xlsx", ".xls", ".csv", ".numbers"}
DOCUMENT_EXTENSIONS = {".pdf", ".docx", ".doc", ".md", ".txt", *SPREADSHEET_EXTENSIONS}


def slugify(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", value.lower().replace("&", "and")).strip("-")


def classify(path: Path) -> str:
    name = path.name.lower()
    if path.suffix.lower() in SPREADSHEET_EXTENSIONS:
        return "spreadsheet"
    if name.startswith("lm_") or name.startswith("lm-") or "lead_magnet" in name:
        return "lead magnet"
    if "90day" in name or "90_day" in name or "90-day" in name or "content_calendar" in name or "social_calendar" in name:
        return "90-day calendar"
    if "ai_playbook" in name or "prompt" in name:
        return "ai playbook"
    if "brand_kit" in name:
        return "brand kit"
    if "email" in name or "communication" in name or "template" in name:
        return "templates"
    if any(word in name for word in ["intake", "onboarding", "questionnaire", "scripts", "system", "tracker", "planner", "calculator", "seo", "membership", "preapproval", "consult", "retention"]):
        return "client system"
    if path.suffix.lower() == ".md":
        return "sales copy"
    return "other"


def pdf_pages(path: Path) -> int | None:
    try:
        return len(PdfReader(str(path)).pages)
    except Exception:
        return None


def spreadsheet_summary(path: Path) -> dict[str, object]:
    summary: dict[str, object] = {"sheets": [], "rows": None, "cols": None}
    if path.suffix.lower() == ".csv":
        try:
            with path.open("r", encoding="utf-8-sig", newline="") as handle:
                reader = csv.reader(handle)
                rows = list(reader)
            summary["rows"] = len(rows)
            summary["cols"] = max((len(row) for row in rows), default=0)
        except Exception as exc:
            summary["error"] = str(exc)
        return summary
    if path.suffix.lower() == ".xlsx" and load_workbook:
        try:
            wb = load_workbook(path, read_only=True, data_only=False)
            summary["sheets"] = wb.sheetnames
            if wb.sheetnames:
                ws = wb[wb.sheetnames[0]]
                summary["rows"] = ws.max_row
                summary["cols"] = ws.max_column
            wb.close()
        except Exception as exc:
            summary["error"] = str(exc)
    return summary


def niche_dirs() -> list[Path]:
    return [
        item
        for item in sorted(SOURCE_ROOT.iterdir())
        if item.is_dir() and item.name not in CUSTOMER_SKIP and not item.name.startswith(".")
    ]


def collect() -> tuple[list[dict[str, object]], list[dict[str, object]]]:
    rows: list[dict[str, object]] = []
    niche_summaries: list[dict[str, object]] = []
    for niche_dir in niche_dirs():
        files = [
            path
            for path in sorted(niche_dir.rglob("*"))
            if path.is_file() and path.suffix.lower() in DOCUMENT_EXTENSIONS and ".DS_Store" not in path.name
        ]
        type_counter: Counter[str] = Counter()
        total_pdf_pages = 0
        spreadsheet_count = 0
        for path in files:
            doc_type = classify(path)
            type_counter[doc_type] += 1
            rel = path.relative_to(SOURCE_ROOT).as_posix()
            page_count = pdf_pages(path) if path.suffix.lower() == ".pdf" else None
            if page_count:
                total_pdf_pages += page_count
            sheet_info = spreadsheet_summary(path) if path.suffix.lower() in SPREADSHEET_EXTENSIONS else {}
            if path.suffix.lower() in SPREADSHEET_EXTENSIONS:
                spreadsheet_count += 1
            rows.append(
                {
                    "niche": niche_dir.name,
                    "path": rel,
                    "filename": path.name,
                    "extension": path.suffix.lower(),
                    "type": doc_type,
                    "pages": page_count,
                    "spreadsheet_sheets": ", ".join(sheet_info.get("sheets", []) or []),
                    "spreadsheet_rows": sheet_info.get("rows"),
                    "spreadsheet_cols": sheet_info.get("cols"),
                    "notes": sheet_info.get("error", ""),
                }
            )

        required_types = ["lead magnet", "ai playbook", "90-day calendar", "brand kit", "templates", "client system"]
        missing = [doc_type for doc_type in required_types if type_counter[doc_type] == 0]
        niche_summaries.append(
            {
                "niche": niche_dir.name,
                "slug": slugify(niche_dir.name),
                "file_count": len(files),
                "pdf_count": sum(1 for item in files if item.suffix.lower() == ".pdf"),
                "spreadsheet_count": spreadsheet_count,
                "total_pdf_pages": total_pdf_pages,
                "document_types": dict(type_counter),
                "missing_expected_types": missing,
            }
        )
    return rows, niche_summaries


def write_outputs(rows: list[dict[str, object]], summaries: list[dict[str, object]]) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    csv_path = OUT_DIR / "product-file-inventory.csv"
    with csv_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()) if rows else [])
        writer.writeheader()
        writer.writerows(rows)

    (OUT_DIR / "product-audit.json").write_text(
        json.dumps({"files": rows, "niches": summaries}, indent=2),
        encoding="utf-8",
    )

    totals = Counter()
    for row in rows:
        totals[row["type"]] += 1
    spreadsheet_rows = [row for row in rows if row["type"] == "spreadsheet"]
    missing_rows = [summary for summary in summaries if summary["missing_expected_types"]]

    md = [
        "# Content Elevated Product Audit",
        "",
        f"- Customer-facing product folders scanned: **{len(summaries)}**",
        f"- Product files found: **{len(rows)}**",
        f"- PDFs found: **{sum(1 for row in rows if row['extension'] == '.pdf')}**",
        f"- Spreadsheets found: **{len(spreadsheet_rows)}**",
        f"- Total source PDF pages: **{sum(summary['total_pdf_pages'] for summary in summaries)}**",
        "",
        "## File Types",
        "",
    ]
    for key, value in sorted(totals.items()):
        md.append(f"- {key}: **{value}**")
    md.extend(["", "## Spreadsheets To Preserve / Rebrand", ""])
    if spreadsheet_rows:
        for row in spreadsheet_rows:
            details = []
            if row.get("spreadsheet_sheets"):
                details.append(f"sheets: {row['spreadsheet_sheets']}")
            if row.get("spreadsheet_rows"):
                details.append(f"{row['spreadsheet_rows']} rows x {row['spreadsheet_cols']} cols")
            md.append(f"- `{row['path']}`" + (f" ({'; '.join(details)})" if details else ""))
    else:
        md.append("- None found.")

    md.extend(["", "## Missing / Odd Expected Types", ""])
    if missing_rows:
        for summary in missing_rows:
            md.append(f"- **{summary['niche']}**: missing {', '.join(summary['missing_expected_types'])}")
    else:
        md.append("- No obvious gaps across the standard bundle structure.")

    md.extend(["", "## Niche Summary", ""])
    for summary in summaries:
        md.append(
            f"- **{summary['niche']}**: {summary['pdf_count']} PDFs, "
            f"{summary['spreadsheet_count']} spreadsheets, {summary['total_pdf_pages']} PDF pages"
        )
    (OUT_DIR / "product-audit.md").write_text("\n".join(md) + "\n", encoding="utf-8")


def main() -> None:
    rows, summaries = collect()
    write_outputs(rows, summaries)
    print(OUT_DIR / "product-audit.md")
    print(OUT_DIR / "product-file-inventory.csv")
    print(f"{len(summaries)} folders, {len(rows)} files")


if __name__ == "__main__":
    main()
